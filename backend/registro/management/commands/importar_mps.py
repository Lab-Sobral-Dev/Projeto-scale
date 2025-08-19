from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from registro.models import MateriaPrima

# Aproveita as mesmas utilidades do arquivo de produtos (copiadas aqui por simplicidade)
TRUE_SET = {"1","true","t","sim","s","y","yes"}
FALSE_SET = {"0","false","f","nao","não","n","no"}

def norm(s):
    return (str(s or "")).strip()

def normalize_key(k: str):
    k = (k or "").strip().lower()
    k = k.replace("ç", "c").replace("á","a").replace("à","a").replace("ã","a").replace("â","a") \
         .replace("é","e").replace("ê","e").replace("í","i").replace("ó","o").replace("ô","o").replace("õ","o") \
         .replace("ú","u").replace("ü","u")
    while "  " in k:
        k = k.replace("  ", " ")
    return k

NORMALIZADORES = {
    "nome": {"nome", "materia-prima", "matéria-prima", "nome da mp", "descricao", "descrição", "mp"},
    "codigo_interno": {
        "codigo_interno","código interno","codigo interno",
        "codigo","código","cod interno","cod","sku","código da mp","codigo da mp"
    },
    "ativo": {"ativo","status","habilitado","situacao","situação","enable","enabled"},
}

def detectar_header(ws, max_busca=10):
    for row_idx in range(1, max_busca + 1):
        possiveis = [normalize_key(c.value) for c in ws[row_idx]]
        if any(p in NORMALIZADORES["nome"] for p in possiveis) and \
           any(p in NORMALIZADORES["codigo_interno"] for p in possiveis):
            return row_idx
    return 1

def mapear_headers(ws, header_row):
    headers = {}
    for idx, cell in enumerate(ws[header_row], start=1):
        headers[normalize_key(cell.value)] = idx

    def achar(set_alias):
        for alias in set_alias:
            if alias in headers:
                return headers[alias]
        return None

    return {
        "nome": achar(NORMALIZADORES["nome"]),
        "codigo_interno": achar(NORMALIZADORES["codigo_interno"]),
        "ativo": achar(NORMALIZADORES["ativo"]),
        "headers_raw": headers,
    }

def to_bool(val, default=True):
    if val is None or str(val).strip() == "":
        return default
    s = str(val).strip().lower()
    if s in TRUE_SET:
        return True
    if s in FALSE_SET:
        return False
    try:
        return float(s) != 0.0
    except Exception:
        return default

class Command(BaseCommand):
    help = "Importa matérias-primas de registro/bases/mp.xlsx para MateriaPrima."

    def add_arguments(self, parser):
        parser.add_argument("--arquivo", default="registro/bases/mp.xlsx", help="Caminho do XLSX")
        parser.add_argument("--sheet", default=None, help="Nome da aba (opcional)")
        parser.add_argument("--header-row", type=int, default=None, help="Linha do cabeçalho (1-based)")
        parser.add_argument("--verbose", action="store_true", help="Logar linhas processadas")
        parser.add_argument("--dry-run", action="store_true", help="Não salvar; apenas simular")

    @transaction.atomic
    def handle(self, *args, **opts):
        arquivo = opts.get("arquivo")
        sheet = opts.get("sheet")
        header_row = opts.get("header_row")  # <-- underscore
        verbose = bool(opts.get("verbose"))
        dry_run = bool(opts.get("dry_run"))

        try:
            wb = load_workbook(filename=arquivo, data_only=True)
        except Exception as e:
            raise CommandError(f"Não foi possível abrir '{arquivo}': {e}")

        ws = wb[sheet] if sheet else wb.active
        header_row = header_row or detectar_header(ws)
        mapa = mapear_headers(ws, header_row)

        if mapa["nome"] is None or mapa["codigo_interno"] is None:
            raise CommandError(
                f"Cabeçalhos não encontrados. Headers lidos: {list(mapa['headers_raw'].keys())}. "
                "Use --sheet/--header-row ou ajuste os cabeçalhos."
            )

        criados = atualizados = ignorados = erros = 0
        start_data_row = header_row + 1

        for i, row in enumerate(ws.iter_rows(min_row=start_data_row, values_only=True), start=start_data_row):
            try:
                nome = norm(row[mapa["nome"] - 1]) if mapa["nome"] else ""
                codigo = norm(row[mapa["codigo_interno"] - 1]) if mapa["codigo_interno"] else ""
                ativo = to_bool(row[mapa["ativo"] - 1]) if mapa["ativo"] else True

                if not nome or not codigo:
                    ignorados += 1
                    if verbose:
                        self.stdout.write(f"[L{i}] ignorado (nome/codigo vazios)")
                    continue

                if dry_run:
                    if verbose:
                        self.stdout.write(f"[L{i}] DRY-RUN -> {codigo} | {nome} | ativo={ativo}")
                    continue

                obj, created = MateriaPrima.objects.get_or_create(
                    codigo_interno=codigo,
                    defaults={"nome": nome, "ativo": ativo},
                )
                if created:
                    criados += 1
                    if verbose:
                        self.stdout.write(f"[L{i}] criado -> {codigo}")
                else:
                    changed = False
                    if obj.nome != nome:
                        obj.nome = nome
                        changed = True
                    if obj.ativo != ativo:
                        obj.ativo = ativo
                        changed = True
                    if changed:
                        obj.save(update_fields=["nome", "ativo"])
                        atualizados += 1
                        if verbose:
                            self.stdout.write(f"[L{i}] atualizado -> {codigo}")
                    else:
                        ignorados += 1
                        if verbose:
                            self.stdout.write(f"[L{i}] sem mudanças -> {codigo}")
            except Exception as e:
                erros += 1
                self.stderr.write(f"[L{i}] ERRO: {e!r}")

        if dry_run:
            self.stdout.write(self.style.NOTICE("DRY-RUN concluído (nada salvo)."))

        self.stdout.write(self.style.SUCCESS(
            f"Matérias-primas -> Criadas: {criados} | Atualizadas: {atualizados} | Ignoradas: {ignorados} | Erros: {erros}"
        ))

