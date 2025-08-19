from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

NORMALIZADORES = {
    "nome": {"nome", "produto", "nome do produto", "descrição", "descricao"},
    "codigo_interno": {
        "codigo_interno", "código interno", "codigo interno",
        "codigo", "código", "cod interno", "cod", "sku", "código do produto"
    },
    "ativo": {"ativo", "status", "habilitado", "situação", "situacao", "enable", "enabled"},
}

def norm(s):
    return (str(s or "")).strip()

def normalize_key(k: str):
    k = (k or "").strip().lower()
    k = k.replace("ç", "c").replace("á","a").replace("à","a").replace("ã","a").replace("â","a") \
         .replace("é","e").replace("ê","e").replace("í","i").replace("ó","o").replace("ô","o").replace("õ","o") \
         .replace("ú","u").replace("ü","u")
    while "  " in k:
        k = k.replace("  ", " ")
    return k

def detectar_header(ws, max_busca=10):
    """Procura uma linha de 1..max_busca que contenha algo parecido com os headers."""
    for row_idx in range(1, max_busca + 1):
        possiveis = [normalize_key(c.value) for c in ws[row_idx]]
        if any(p in NORMALIZADORES["nome"] for p in possiveis) and \
           any(p in NORMALIZADORES["codigo_interno"] for p in possiveis):
            return row_idx
    return 1  # fallback

def mapear_headers(ws, header_row):
    headers = {}
    for idx, cell in enumerate(ws[header_row], start=1):
        k = normalize_key(cell.value)
        headers[k] = idx

    def achar(*aliases):
        for alias in aliases:
            # alias já normalizado
            if alias in headers:
                return headers[alias]
        return None

    # montar mapa com sinônimos
    col_nome = achar(*NORMALIZADORES["nome"])
    col_codigo = achar(*NORMALIZADORES["codigo_interno"])
    col_ativo = achar(*NORMALIZADORES["ativo"])
    return {"nome": col_nome, "codigo_interno": col_codigo, "ativo": col_ativo, "headers_raw": headers}

class Command(BaseCommand):
    help = "Inspeciona uma planilha XLSX: lista abas, detecta cabeçalho e mostra amostra de linhas parseadas."

    def add_arguments(self, parser):
        parser.add_argument("--arquivo", required=True, help="Caminho do XLSX")
        parser.add_argument("--sheet", default=None, help="Nome da aba (opcional)")
        parser.add_argument("--header-row", type=int, default=None, help="Linha do cabeçalho (1-based)")

    def handle(self, *args, **opts):
        arquivo = opts.get("arquivo")
        sheet = opts.get("sheet")
        header_row_cli = opts.get("header_row")  # <-- underscore

        try:
            wb = load_workbook(filename=arquivo, data_only=True)
        except Exception as e:
            raise CommandError(f"Não foi possível abrir '{arquivo}': {e}")

        self.stdout.write(self.style.NOTICE(f"Arquivo: {arquivo}"))
        self.stdout.write(f"Abas: {', '.join(wb.sheetnames)}")

        ws = wb[sheet] if sheet else wb.active
        self.stdout.write(f"Aba usada: {ws.title}")

        header_row = header_row_cli or detectar_header(ws)
        self.stdout.write(f"Cabeçalho detectado na linha: {header_row}")

        mapa = mapear_headers(ws, header_row)
        self.stdout.write(f"Mapeamento -> nome: {mapa['nome']}, codigo_interno: {mapa['codigo_interno']}, ativo: {mapa['ativo']}")
        self.stdout.write(f"Headers (normalizados): {list(mapa['headers_raw'].keys())[:30]}")

        self.stdout.write(self.style.NOTICE("Amostra (5 linhas após o cabeçalho):"))
        count = 0
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if count >= 5:
                break
            nome = row[mapa['nome'] - 1] if mapa['nome'] else None
            codigo = row[mapa['codigo_interno'] - 1] if mapa['codigo_interno'] else None
            ativo = row[mapa['ativo'] - 1] if mapa['ativo'] else None
            self.stdout.write(f"- nome={nome!r} | codigo_interno={codigo!r} | ativo={ativo!r}")
            count += 1

        if mapa["nome"] is None or mapa["codigo_interno"] is None:
            raise CommandError("ERRO: Não foi possível mapear 'nome' e/ou 'codigo_interno'. Ajuste o cabeçalho ou use --header-row/--sheet.")
